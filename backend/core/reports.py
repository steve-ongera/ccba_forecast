import os
from datetime import datetime

from django.conf import settings as django_settings

from .models import Sale, Forecast, Recommendation


REPORTS_DIR = os.path.join(django_settings.MEDIA_ROOT, "reports")


def _ensure_reports_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _apply_filters(queryset, filters, date_field):
    product_id = filters.get("product_id")
    region_id = filters.get("region_id")
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")

    if product_id:
        queryset = queryset.filter(product_id=product_id)
    if region_id:
        queryset = queryset.filter(region_id=region_id)
    if date_from:
        queryset = queryset.filter(**{f"{date_field}__gte": date_from})
    if date_to:
        queryset = queryset.filter(**{f"{date_field}__lte": date_to})
    return queryset


def _gather_data(filters):
    sales = _apply_filters(
        Sale.objects.select_related("product", "region"), filters, "sales_date"
    ).order_by("-sales_date")[:500]

    forecasts = _apply_filters(
        Forecast.objects.select_related("product", "region", "model_used"), filters, "forecast_date"
    ).order_by("-forecast_date")[:500]

    recommendations = Recommendation.objects.select_related("product", "region").order_by("-created_at")[:50]

    return sales, forecasts, recommendations


def build_report(report_format, filters):
    """
    Builds a PDF or Excel report based on filters:
    {product_id, region_id, date_from, date_to}
    Returns a path (relative to MEDIA_ROOT) to the generated file.
    """
    _ensure_reports_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sales, forecasts, recommendations = _gather_data(filters or {})

    if report_format == "excel":
        filename = f"demand_report_{timestamp}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        _build_excel(filepath, sales, forecasts, recommendations)
    else:
        filename = f"demand_report_{timestamp}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        _build_pdf(filepath, sales, forecasts, recommendations)

    return f"reports/{filename}"


def _build_excel(filepath, sales, forecasts, recommendations):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC0019", end_color="DC0019", fill_type="solid")

    def write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws_sales = wb.active
    ws_sales.title = "Sales"
    write_sheet(
        ws_sales,
        ["Date", "Product", "Region", "Qty Sold", "Unit Price", "Revenue", "Promotion"],
        [
            [
                s.sales_date, s.product.name, s.region.name, s.quantity_sold,
                float(s.unit_price), float(s.revenue), "Yes" if s.promotion_active else "No",
            ]
            for s in sales
        ],
    )

    ws_forecast = wb.create_sheet("Forecasts")
    write_sheet(
        ws_forecast,
        ["Forecast Date", "Product", "Region", "Horizon", "Predicted Demand", "Actual Demand", "Algorithm"],
        [
            [
                f.forecast_date, f.product.name, f.region.name, f.get_horizon_display(),
                float(f.predicted_demand), float(f.actual_demand) if f.actual_demand else None,
                f.model_used.get_algorithm_display() if f.model_used else "N/A",
            ]
            for f in forecasts
        ],
    )

    ws_recs = wb.create_sheet("Recommendations")
    write_sheet(
        ws_recs,
        ["Date", "Product", "Region", "Severity", "Change %", "Message"],
        [
            [r.created_at.strftime("%Y-%m-%d %H:%M"), r.product.name, r.region.name,
             r.get_severity_display(), r.change_pct, r.message]
            for r in recommendations
        ],
    )

    wb.save(filepath)


def _build_pdf(filepath, sales, forecasts, recommendations):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("CCBA Demand Forecast Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    def add_table(title, headers, rows):
        elements.append(Paragraph(title, styles["Heading2"]))
        data = [headers] + rows
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DC0019")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 16))

    add_table(
        "Sales (latest)",
        ["Date", "Product", "Region", "Qty", "Unit Price", "Revenue"],
        [
            [str(s.sales_date), s.product.name, s.region.name, str(s.quantity_sold),
             f"{s.unit_price:.2f}", f"{s.revenue:.2f}"]
            for s in sales[:50]
        ],
    )

    add_table(
        "Forecasts (latest)",
        ["Date", "Product", "Region", "Horizon", "Predicted", "Actual"],
        [
            [str(f.forecast_date), f.product.name, f.region.name, f.get_horizon_display(),
             f"{f.predicted_demand:.2f}", f"{f.actual_demand:.2f}" if f.actual_demand else "-"]
            for f in forecasts[:50]
        ],
    )

    add_table(
        "AI Recommendations",
        ["Date", "Product", "Region", "Severity", "Change %", "Message"],
        [
            [r.created_at.strftime("%Y-%m-%d"), r.product.name, r.region.name,
             r.get_severity_display(), f"{r.change_pct:+.1f}%", r.message[:80]]
            for r in recommendations[:30]
        ],
    )

    doc.build(elements)